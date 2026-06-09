from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Classificação CFOPs"

COR_HEADER_BG = "0D2137"
COR_SECAO_BG  = "D6E4F7"
COR_SECAO_FNT = "1F3864"
COR_BRANCO    = "FFFFFF"

thin = Side(style='thin', color="CCCCCC")
borda = Border(left=thin, right=thin, top=thin, bottom=thin)

COR_CLASS = {
    "Venda":              "C6EFCE",
    "Devolucao saida":    "FCE4D6",
    "Remessa":            "FFF2CC",
    "Retorno de remessa": "E2CFFF",
    "Outros":             "F2F2F2",
    "Compra":             "DDEEFF",
    "Entrada de remessa": "D9F0FF",
    "Devolucao venda":    "FFD7D7",
    "Uso Consumo":        "EBEBEB",
    "Ativo Imobilizado":  "E8E0FF",
}

# Título
ws.row_dimensions[1].height = 28
ws.merge_cells("A1:E1")
c = ws["A1"]
c.value = "Classificação de CFOPs — Sistema Cúmplice / KDM Confecções"
c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
c.fill = PatternFill("solid", fgColor=COR_HEADER_BG)
c.alignment = Alignment(horizontal="center", vertical="center")

ws.row_dimensions[2].height = 14
ws.merge_cells("A2:E2")
c = ws["A2"]
c.value = "Validado contra SPED EFD ICMS/IPI — Abril/2026  ·  Cruzamento Domínio × SPED"
c.font = Font(name="Arial", size=8, color="888888", italic=True)
c.alignment = Alignment(horizontal="center", vertical="center")

# Cabeçalho colunas
ws.row_dimensions[3].height = 22
headers = ["CFOP", "Classificação", "Descrição", "Entrada/Saída", "Observação"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=col, value=h)
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c.fill = PatternFill("solid", fgColor=COR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = borda

def add_secao(ws, row, texto):
    ws.row_dimensions[row].height = 16
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1, value=texto)
    c.font = Font(name="Arial", bold=True, color=COR_SECAO_FNT, size=9)
    c.fill = PatternFill("solid", fgColor=COR_SECAO_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = borda

def add_row(ws, row, cfop, classe, desc, tipo, obs, cor_key):
    ws.row_dimensions[row].height = 16
    bg = COR_CLASS.get(cor_key, COR_BRANCO)
    vals = [cfop, classe, desc, tipo, obs]
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Arial", size=9, bold=(col == 2))
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(vertical="center", wrap_text=True,
                                horizontal="center" if col in (1, 2, 4) else "left")
        c.border = borda

r = 4

add_secao(ws, r, "SAÍDAS"); r += 1

saidas = [
    ("5101 / 5102", "Venda", "Venda de produto industrializado / mercadoria adquirida de terceiros", "Saída", "", "Venda"),
    ("5103 – 5106", "Venda", "Venda fora do estab. / ao Governo / a não contribuinte", "Saída", "", "Venda"),
    ("6101 / 6102", "Venda", "Venda interestadual de produto / mercadoria", "Saída", "", "Venda"),
    ("6103 – 6106", "Venda", "Venda interestadual (fora estab. / Governo / não contrib.)", "Saída", "", "Venda"),
    ("6107", "Venda", "Venda p/ Zona Franca de Manaus / ALC", "Saída", "", "Venda"),
    ("6108", "Venda", "Venda interestadual c/ retenção ICMS-ST", "Saída", "", "Venda"),
    ("5401 / 5403 / 5405", "Venda", "Venda com substituição tributária (ICMS-ST)", "Saída", "", "Venda"),
    ("6401 / 6403", "Venda", "Venda interestadual c/ ICMS-ST", "Saída", "", "Venda"),
    ("5933 / 6933", "Venda", "Prestação de serviço (ISSQN — NFS-e)", "Saída", "Tributado pelo município", "Venda"),
    ("5124 / 6124", "Venda", "Industrialização efetuada p/ encomendante", "Saída", "É receita de serviço prestado", "Venda"),
    ("5201 / 5202", "Devolução (saída)", "Devolução de compra p/ industrialização / mercadoria", "Saída", "Deduz saldo de compras", "Devolucao saida"),
    ("6201 / 6202", "Devolução (saída)", "Devolução interestadual de compra", "Saída", "", "Devolucao saida"),
    ("5901 / 6901", "Remessa", "Remessa p/ industrialização por encomenda", "Saída", "Sem transferência de propriedade", "Remessa"),
    ("5903 / 6903", "Remessa", "Remessa p/ venda fora do estabelecimento", "Saída", "", "Remessa"),
    ("5910", "Remessa", "Remessa em bonificação", "Saída", "", "Remessa"),
    ("6911", "Remessa", "Remessa p/ armazenagem / depósito", "Saída", "", "Remessa"),
    ("6912", "Remessa", "Remessa p/ demonstração ou mostruário", "Saída", "", "Remessa"),
    ("6908", "Remessa", "Retorno de mercadoria depositada em terceiros", "Saída", "", "Remessa"),
    ("5902 / 6902", "Retorno de remessa", "Retorno de industrialização ao encomendante", "Saída", "", "Retorno de remessa"),
    ("5904", "Retorno de remessa", "Retorno de remessa p/ venda fora do estab.", "Saída", "", "Retorno de remessa"),
    ("6913", "Retorno de remessa", "Retorno de demonstração ou mostruário", "Saída", "", "Retorno de remessa"),
    ("6923", "Retorno de remessa", "Retorno de depósito fechado", "Saída", "", "Retorno de remessa"),
    ("5929", "Retorno de remessa", "Lançamento de crédito ICMS acumulado — Conv. 29/90", "Saída", "VL_DOC = 0 (simbólico) · COD_SIT=08", "Retorno de remessa"),
    ("5949 / 6949", "Outros", "Outra saída de mercadoria ou serviço", "Saída", "Validado: inclui veículo R$235k (NF 17094 — WAY)", "Outros"),
]
for row_data in saidas:
    add_row(ws, r, *row_data); r += 1

add_secao(ws, r, "ENTRADAS"); r += 1

entradas = [
    ("1101 / 2101", "Compra", "Compra p/ industrialização", "Entrada", "", "Compra"),
    ("1102 / 2102", "Compra", "Compra p/ comercialização", "Entrada", "", "Compra"),
    ("1124 / 2124", "Compra", "Compra p/ industrialização (outra modalidade)", "Entrada", "", "Compra"),
    ("1902 / 2902", "Retorno de remessa", "Retorno de mercadoria enviada p/ industrialização", "Entrada", "Soma com 2923 = R$1.004.091,06 (validado abr/2026)", "Retorno de remessa"),
    ("2923", "Retorno de remessa", "Entrada do vendedor remetente em venda à ordem", "Entrada", "Fecha ciclo de consignação industrial", "Retorno de remessa"),
    ("1901 / 2901", "Entrada de remessa", "Entrada p/ industrialização por encomenda", "Entrada", "Recebimento temporário", "Entrada de remessa"),
    ("2216", "Entrada de remessa", "Entrada interestadual de mercadoria p/ industrialização", "Entrada", "", "Entrada de remessa"),
    ("2911", "Entrada de remessa", "Retorno de remessa p/ venda fora do estab.", "Entrada", "", "Entrada de remessa"),
    ("2912", "Entrada de remessa", "Entrada p/ demonstração ou mostruário", "Entrada", "", "Entrada de remessa"),
    ("1201 / 1202", "Devolução de venda", "Devolução de venda de produto próprio / mercadoria", "Entrada", "Deduz faturamento", "Devolucao venda"),
    ("2201 / 2202", "Devolução de venda", "Devolução interestadual de venda", "Entrada", "", "Devolucao venda"),
    ("1203 / 1204", "Devolução de venda", "Devolução de venda p/ ZFM/ALC", "Entrada", "", "Devolucao venda"),
    ("1407", "Uso / Consumo", "Compra de mercadoria p/ uso e consumo c/ ICMS-ST", "Entrada", "Não compõe custo de produto", "Uso Consumo"),
    ("1556 / 2556", "Uso / Consumo", "Compra de material p/ uso ou consumo", "Entrada", "", "Uso Consumo"),
    ("1551 / 2551", "Ativo Imobilizado", "Aquisição de bem p/ ativo permanente", "Entrada", "", "Ativo Imobilizado"),
]
for row_data in entradas:
    add_row(ws, r, *row_data); r += 1

ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 54
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 44
ws.freeze_panes = "A4"

# Aba Legenda
wl = wb.create_sheet("Legenda")
wl.column_dimensions["A"].width = 22
wl.column_dimensions["B"].width = 54

legenda = [
    ("Classificação", "Significado"),
    ("Venda", "Receita real de venda de produto ou serviço prestado"),
    ("Compra", "Entrada de insumo ou mercadoria p/ produção ou comercialização"),
    ("Remessa", "Movimentação de estoque sem transferência de propriedade — NÃO é receita"),
    ("Retorno de remessa", "Retorno físico de mercadoria enviada em remessa — entra no saldo fiscal"),
    ("Entrada de remessa", "Recebimento temporário de mercadoria de terceiros p/ industrialização"),
    ("Devolução (saída)", "Devolução de compra emitida pelo estabelecimento — deduz saldo de compras"),
    ("Devolução de venda", "Devolução recebida de cliente — deduz faturamento"),
    ("Uso / Consumo", "Material auxiliar — não compõe custo de produto"),
    ("Ativo Imobilizado", "Aquisição de bem permanente — ativo fixo"),
    ("Outros", "Operações diversas não enquadradas acima — analisar caso a caso"),
]

cor_map = {
    "Venda": "C6EFCE", "Compra": "DDEEFF", "Remessa": "FFF2CC",
    "Retorno de remessa": "E2CFFF", "Entrada de remessa": "D9F0FF",
    "Devolução (saída)": "FCE4D6", "Devolução de venda": "FFD7D7",
    "Uso / Consumo": "EBEBEB", "Ativo Imobilizado": "E8E0FF", "Outros": "F2F2F2",
}

for i, (k, v) in enumerate(legenda, 1):
    ck = wl.cell(row=i, column=1, value=k)
    cv = wl.cell(row=i, column=2, value=v)
    if i == 1:
        for c in (ck, cv):
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill("solid", fgColor=COR_HEADER_BG)
            c.alignment = Alignment(horizontal="center", vertical="center")
    else:
        bg = cor_map.get(k, COR_BRANCO)
        ck.font = Font(name="Arial", size=9, bold=True)
        cv.font = Font(name="Arial", size=9)
        for c in (ck, cv):
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(vertical="center")
    for c in (ck, cv):
        c.border = borda
    wl.row_dimensions[i].height = 16

out = r"C:\Users\louri\Documents\Projetos\KDM\cumplice-kdm\Classificacao_CFOPs_KDM.xlsx"
wb.save(out)
print("Salvo:", out)
